"""
MEBU Experiment Database Extractor
Reads raw experiment Excel files and populates MEBU_Experiment_Database_v1.xlsx
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os

EXPERIMENTS = [
    {
        "exp_id": "EXP-001",
        "exp_name": "Acceptance Test Dec24",
        "exp_type": "Acceptance Test",
        "file": "01_Master_MEBU_Acceptance_Test1824Dec24_rev6Jan25_LAB_UPDATED.xlsx",
        "sheet": "Master Template",
        "start_date": "18-Dec-24",
        "op_date_row": 3,
        "lab_date_row": 4,
        "day_row": 5,
        "data_col_start": 3,
        "params": {
            "CrkConv":     {"row": 93,  "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 78},
            "CrkConv_Corr":{"row": 94,  "unit": "wt%", "category": "Conversion",         "art_low": None, "art_high": None},
            "NiConv":      {"row": 98,  "unit": "wt%", "category": "Metal Conversion",   "art_low": 88, "art_high": 100},
            "VConv":       {"row": 99,  "unit": "wt%", "category": "Metal Conversion",   "art_low": 98, "art_high": 100},
            "NiV_Conv":    {"row": 103, "unit": "wt%", "category": "Metal Conversion",   "art_low": 96, "art_high": 100},
            "FeConv":      {"row": 105, "unit": "wt%", "category": "Metal Conversion",   "art_low": None, "art_high": None},
            "SConv":       {"row": 107, "unit": "wt%", "category": "Impurity Conversion","art_low": 80, "art_high": 100},
            "NConv":       {"row": 110, "unit": "wt%", "category": "Impurity Conversion","art_low": None, "art_high": None},
            "MCRConv":     {"row": 120, "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 100},
            "Sedimentation":{"row": 161,"unit": "ppm", "category": "Sedimentation",      "art_low": 0, "art_high": 1500},
        },
    },
    {
        "exp_id": "EXP-002",
        "exp_name": "R1R2 Repro Nov/Dec25",
        "exp_type": "R1R2+Repro",
        "file": "01_Master_MEBU_Result_R1R2_with_ReproNovDec25.xlsx",
        "sheet": "Master Template",
        "start_date": "01-Nov-25",
        "op_date_row": 3,
        "lab_date_row": 4,
        "day_row": 5,
        "data_col_start": 5,
        "params": {
            "CrkConv":      {"row": 141, "unit": "wt%", "category": "Conversion",         "art_low": 36, "art_high": 70},
            "CrkConv_Corr": {"row": 142, "unit": "wt%", "category": "Conversion",         "art_low": None, "art_high": None},
            "NiConv":       {"row": 149, "unit": "wt%", "category": "Metal Conversion",   "art_low": 88, "art_high": 100},
            "VConv":        {"row": 153, "unit": "wt%", "category": "Metal Conversion",   "art_low": 98, "art_high": 100},
            "NiV_Conv":     {"row": 164, "unit": "wt%", "category": "Metal Conversion",   "art_low": 96, "art_high": 100},
            "FeConv":       {"row": 185, "unit": "wt%", "category": "Metal Conversion",   "art_low": None, "art_high": None},
            "SConv":        {"row": 188, "unit": "wt%", "category": "Impurity Conversion","art_low": 80, "art_high": 100},
            "NConv":        {"row": 191, "unit": "wt%", "category": "Impurity Conversion","art_low": None, "art_high": None},
            "MCRConv":      {"row": 201, "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 100},
            "Sedimentation":{"row": 267, "unit": "ppm", "category": "Sedimentation",      "art_low": 0, "art_high": 1500},
        },
    },
    {
        "exp_id": "EXP-003",
        "exp_name": "LnH Sedim Jul/Aug25",
        "exp_type": "L&H + Sedim",
        "file": "01_Master_MEBU_Result_LnH_Sedim_Include_N2_GasJulAug25.xlsx",
        "sheet": "Master Template",
        "start_date": "01-Jul-25",
        "op_date_row": 3,
        "lab_date_row": 4,
        "day_row": 5,
        "data_col_start": 5,
        "params": {
            "CrkConv":      {"row": 139, "unit": "wt%", "category": "Conversion",         "art_low": 36, "art_high": 70},
            "CrkConv_Corr": {"row": 140, "unit": "wt%", "category": "Conversion",         "art_low": None, "art_high": None},
            "NiConv":       {"row": 147, "unit": "wt%", "category": "Metal Conversion",   "art_low": 88, "art_high": 100},
            "VConv":        {"row": 151, "unit": "wt%", "category": "Metal Conversion",   "art_low": 98, "art_high": 100},
            "NiV_Conv":     {"row": 162, "unit": "wt%", "category": "Metal Conversion",   "art_low": 96, "art_high": 100},
            "FeConv":       {"row": 182, "unit": "wt%", "category": "Metal Conversion",   "art_low": None, "art_high": None},
            "SConv":        {"row": 185, "unit": "wt%", "category": "Impurity Conversion","art_low": 80, "art_high": 100},
            "NConv":        {"row": 188, "unit": "wt%", "category": "Impurity Conversion","art_low": None, "art_high": None},
            "MCRConv":      {"row": 198, "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 100},
            "Sedimentation":{"row": 264, "unit": "ppm", "category": "Sedimentation",      "art_low": 0, "art_high": 1500},
        },
    },
    {
        "exp_id": "EXP-004",
        "exp_name": "Repeat Accept Test Feb/Mar26",
        "exp_type": "Acceptance Test",
        "file": "01_Master_MEBU_Result_Repeat_Acceptance_Test_FebMar26.xlsx",
        "sheet": "Master Template",
        "start_date": "23-Feb-26",
        "op_date_row": 3,
        "lab_date_row": 4,
        "day_row": 5,
        "data_col_start": 5,
        "params": {
            "CrkConv":      {"row": 155, "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 78},
            "CrkConv_Corr": {"row": 156, "unit": "wt%", "category": "Conversion",         "art_low": None, "art_high": None},
            "NiConv":       {"row": 165, "unit": "wt%", "category": "Metal Conversion",   "art_low": 88, "art_high": 100},
            "VConv":        {"row": 171, "unit": "wt%", "category": "Metal Conversion",   "art_low": 98, "art_high": 100},
            "NiV_Conv":     {"row": 184, "unit": "wt%", "category": "Metal Conversion",   "art_low": 96, "art_high": 100},
            "FeConv":       {"row": 207, "unit": "wt%", "category": "Metal Conversion",   "art_low": None, "art_high": None},
            "SConv":        {"row": 210, "unit": "wt%", "category": "Impurity Conversion","art_low": 80, "art_high": 100},
            "NConv":        {"row": 217, "unit": "wt%", "category": "Impurity Conversion","art_low": None, "art_high": None},
            "MCRConv":      {"row": 229, "unit": "wt%", "category": "Conversion",         "art_low": 70, "art_high": 100},
            "Sedimentation":{"row": 299, "unit": "ppm", "category": "Sedimentation",      "art_low": 0, "art_high": 1500},
        },
    },
]


def extract_experiment_data(exp_config, base_dir="."):
    fpath = os.path.join(base_dir, exp_config["file"])
    if not os.path.exists(fpath):
        print(f"  WARNING: File not found: {fpath}")
        return []

    print(f"  Reading: {exp_config['file']}")
    df = pd.read_excel(fpath, sheet_name=exp_config["sheet"], header=None)

    col_start = exp_config["data_col_start"]
    row2 = df.iloc[2]
    section_start_cols = sorted([i for i, v in enumerate(row2) if pd.notna(v) and str(v).strip()])

    # Find the section that starts at or near col_start
    if col_start in section_start_cols:
        idx = section_start_cols.index(col_start)
    else:
        candidates = [c for c in section_start_cols if c <= col_start]
        this_sec = max(candidates) if candidates else col_start
        idx = section_start_cols.index(this_sec) if this_sec in section_start_cols else 0

    if idx + 1 < len(section_start_cols):
        next_section = section_start_cols[idx + 1]
    else:
        next_section = df.shape[1]

    section_cols = list(range(col_start, next_section))

    op_dates = df.iloc[exp_config["op_date_row"], section_cols].values
    lab_dates = df.iloc[exp_config["lab_date_row"], section_cols].values
    days = df.iloc[exp_config["day_row"], section_cols].values

    valid_cols = []
    for i, (c, d, l, dy) in enumerate(zip(section_cols, op_dates, lab_dates, days)):
        try:
            day_num = int(float(dy))
            if day_num < 1:
                continue
        except (ValueError, TypeError):
            continue
        op_d = pd.to_datetime(d, errors="coerce")
        lab_d = pd.to_datetime(l, errors="coerce")
        valid_cols.append({"col": c, "day": day_num, "op_date": op_d, "lab_date": lab_d})

    if not valid_cols:
        print(f"  WARNING: No valid day columns found")
        return []

    print(f"  Found {len(valid_cols)} day-columns (Day {valid_cols[0]['day']} to Day {valid_cols[-1]['day']})")

    records = []
    for col_info in valid_cols:
        c = col_info["col"]
        day = col_info["day"]
        op_d = col_info["op_date"]
        lab_d = col_info["lab_date"]
        op_str = op_d.strftime("%d-%b-%y") if pd.notna(op_d) else ""
        lab_str = lab_d.strftime("%d-%b-%y") if pd.notna(lab_d) else ""

        for param_name, pcfg in exp_config["params"].items():
            try:
                value = df.iloc[pcfg["row"], c]
                if not pd.notna(value):
                    continue
                val_num = float(value)
                # Skip zero for conversions (startup artefact)
                if val_num == 0 and param_name != "Sedimentation":
                    continue
                art_low = pcfg["art_low"]
                art_high = pcfg["art_high"]
                if art_low is not None and art_high is not None:
                    within_spec = "YES" if art_low <= val_num <= art_high else "NO"
                else:
                    within_spec = "N/A"
                records.append({
                    "Exp_ID": exp_config["exp_id"],
                    "Experiment Name": exp_config["exp_name"],
                    "Type": exp_config["exp_type"],
                    "Start Date": exp_config["start_date"],
                    "Day": day,
                    "Op. Date": op_str,
                    "Lab Date": lab_str,
                    "Category": pcfg["category"],
                    "Parameter": param_name,
                    "Unit": pcfg["unit"],
                    "Value": round(val_num, 4),
                    "Value (LHSV Corr.)": "",
                    "ART Ref (Low)": art_low if art_low is not None else "",
                    "ART Ref (High)": art_high if art_high is not None else "",
                    "Within Spec?": within_spec,
                    "Notes": "",
                })
            except (IndexError, ValueError, TypeError):
                continue

    print(f"  Extracted {len(records)} records")
    return records


def write_to_database(all_records, db_path):
    print(f"\nWriting to: {db_path}")
    wb = load_workbook(db_path)
    ws = wb["Database"]

    # Unmerge all merged cells in data area first
    data_start_row = 4
    header_row = 3
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Clear notice rows and find last real data row
    last_row = header_row
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            if cell.value is not None:
                if "▼" in str(cell.value):
                    cell.value = None
                else:
                    last_row = max(last_row, cell.row)
                break

    # Check existing records to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row[0] and str(row[0]).startswith("EXP-"):
            existing.add((str(row[0]), str(row[4]), str(row[8])))

    new_records = [r for r in all_records
                   if (r["Exp_ID"], str(r["Day"]), r["Parameter"]) not in existing]

    print(f"  Existing records: {len(existing)}")
    print(f"  New records to add: {len(new_records)}")

    if not new_records:
        print("  Nothing new to add.")
        wb.save(db_path)
        return

    columns = [
        "Exp_ID", "Experiment Name", "Type", "Start Date", "Day",
        "Op. Date", "Lab Date", "Category", "Parameter", "Unit",
        "Value", "Value (LHSV Corr.", "ART Ref (Low)", "ART Ref (High)",
        "Within Spec?", "Notes"
    ]
    col_map = {
        "Value (LHSV Corr.": "Value (LHSV Corr.)",
    }

    thin = Side(border_style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill = PatternFill("solid", start_color="EBF5FB")
    odd_fill  = PatternFill("solid", start_color="FFFFFF")
    yes_fill  = PatternFill("solid", start_color="D5F5E3")
    no_fill   = PatternFill("solid", start_color="FADBD8")

    write_start = last_row + 1
    row_num = write_start

    for i, rec in enumerate(new_records):
        base_fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx in range(1, 17):
            col_name_key = [
                "Exp_ID", "Experiment Name", "Type", "Start Date", "Day",
                "Op. Date", "Lab Date", "Category", "Parameter", "Unit",
                "Value", "Value (LHSV Corr.)", "ART Ref (Low)", "ART Ref (High)",
                "Within Spec?", "Notes"
            ][col_idx - 1]
            val = rec.get(col_name_key, "")
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_name_key == "Within Spec?":
                cell.fill = yes_fill if val == "YES" else (no_fill if val == "NO" else base_fill)
            else:
                cell.fill = base_fill
            if col_name_key in ("Value",) and isinstance(val, (int, float)):
                cell.number_format = "0.00"
        row_num += 1

    # Re-add notice row (not merged this time – it was an issue)
    ws.cell(row=row_num, column=1,
            value="▼  Archived experiment data continues below...  (rows appended by Archive process)"
            ).font = Font(italic=True, color="808080")

    wb.save(db_path)
    print(f"  DONE: Wrote {len(new_records)} records (rows {write_start} to {row_num-1})")


def main(base_dir=".", db_path=None, exp_ids=None):
    if db_path is None:
        db_path = os.path.join(base_dir, "MEBU_Experiment_Database_v1.xlsx")
    if not os.path.exists(db_path):
        print(f"ERROR: Database file not found: {db_path}")
        sys.exit(1)

    exps = EXPERIMENTS if not exp_ids else [e for e in EXPERIMENTS if e["exp_id"] in exp_ids]
    all_records = []
    for exp in exps:
        print(f"\nProcessing {exp['exp_id']}: {exp['exp_name']}")
        all_records.extend(extract_experiment_data(exp, base_dir=base_dir))

    print(f"\nTotal records extracted: {len(all_records)}")
    if all_records:
        write_to_database(all_records, db_path)
        print("Done! Database updated.")
    else:
        print("No records extracted.")


if __name__ == "__main__":
    base_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    db_path  = sys.argv[2] if len(sys.argv) > 2 else None
    exp_ids  = sys.argv[3:] if len(sys.argv) > 3 else None
    main(base_dir=base_dir, db_path=db_path, exp_ids=exp_ids)
